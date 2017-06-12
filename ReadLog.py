import sys
from openpyxl import load_workbook
input_file = sys.argv[1]
infile = open(input_file, "r")
wb = load_workbook('Template.xlsm')
ws = wb.active
current_row = 1
count = 0
flag = True
trace_found = False
for line in infile:
    if "MSCi" in line:
        next(infile)
        line = next(infile)
        if "MSC OBSERVATION" in line:
            flag = True
        elif "MSC SMS OBSERVATION" in line:
            flag = False
        next(infile)
        next(infile)
        next(infile)
        line = next(infile)
        line_split = line.split(':')
        # checking the traced IMSi
        if line_split[1].strip() == sys.argv[2]:
            trace_found = True
            current_row = current_row + 1
            count = count + 1
            ws.cell(row=current_row, column=1).value = count
        else:
            trace_found = False

    elif trace_found:
        # MSC OBSERVATION REPORT
        if flag:
            if "RECORDING ENTITY" in line:
                line_split = line.split(':')
                ws.cell(row=current_row, column=2).value = int(line_split[1])
            elif "CALL ID" in line:
                line_split = line.split(':')
                ws.cell(row=current_row, column=3).value = line_split[1].strip()
            elif "CALL START" in line:
                line_split = line.split(' ')
                ws.cell(row=current_row, column=4).value = (line_split[12] + ' ' + line_split[14]).strip()
                ws.cell(row=current_row, column=10).value = line_split[19].strip()
            elif "SIGNALLING COMPLETE" in line:
                line_split = line.split(' ')
                ws.cell(row=current_row, column=5).value = (line_split[3] + ' ' + line_split[5]).strip()
                ws.cell(row=current_row, column=11).value = line_split[10].strip()
            elif "ANSWER" in line:
                line_split = line.split(' ')
                ws.cell(row=current_row, column=6).value = (line_split[15] + ' ' + line_split[17]).strip()
                ws.cell(row=current_row, column=12).value = line_split[22].strip()
            elif "CHARGING END" in line:
                line_split = line.split(' ')
                ws.cell(row=current_row, column=7).value = (line_split[10] + ' ' + line_split[12]).strip()
                ws.cell(row=current_row, column=13).value = (line_split[17] + ' ' + line_split[18] + ' ' +
                                                             line_split[19]).strip()
            elif "PAGING TIME" in line:
                line_split = line.split(':')
                line_split_sub = line_split[1].split(' ')
                ws.cell(row=current_row, column=8).value = line_split_sub[1].strip()
                ws.cell(row=current_row, column=14).value = line_split[2].strip()
            elif "EXT CLEAR CODE" in line:
                line_split = line.split(':')
                line_split_sub = line_split[1].split(' ')
                ws.cell(row=current_row, column=9).value = line_split_sub[1].strip()
                ws.cell(row=current_row, column=15).value = line_split[2].strip()
            elif "EXTERNAL FORWARDING COUNTER" in line:
                line_split = line.split(':')
                ws.cell(row=current_row, column=16).value = line_split[1].strip()
            elif "CALLING NUMBER" in line:
                line_split = line.split(':')
                ws.cell(row=current_row, column=17).value = line_split[1].strip()
            elif "CALLED NUMBER" in line:
                line_split = line.split(':')
                ws.cell(row=current_row, column=18).value = line_split[1].strip()
            elif "OUT PULSED NUMBER" in line:
                line_split = line.split(':')
                ws.cell(row=current_row, column=19).value = line_split[1].strip()
            elif "CONNECTED NUMBER" in line:
                line_split = line.split(':')
                ws.cell(row=current_row, column=20).value = line_split[1].strip()
            elif "ROAMING NUMBER" in line:
                line_split = line.split(':')
                ws.cell(row=current_row, column=21).value = line_split[1].strip()
            elif "ADDRESS NUMBER" in line:
                line_split = line.split(':')
                ws.cell(row=current_row, column=22).value = line_split[1].strip()
            # IMSI SUB B IMSi Not Ready
            elif "IMSI" in line:
                line_split = line.split(':')
                line_split_sub = line_split[1].split(' ')
                ws.cell(row=current_row, column=23).value = int(line_split_sub[1])
            # CGR/BSC/PCM-TSL
            elif "CGR/BSC/PCM-TSL" in line:
                line_split = line.split(':')
            # MCC / MNC
            elif "MCC/MNC" in line:
                line_split = line.split(':')
                line_split_sub = line_split[1].split('           ')
                ws.cell(row=current_row, column=25).value = line_split_sub[0].strip()
                ws.cell(row=current_row, column=38).value = line_split_sub[1].strip()
            elif "LAC/CI/CELL BAND" in line:
                line_split = line.split(':')
                line_split_sub = line_split[1].split('      ')
                ws.cell(row=current_row, column=26).value = line_split_sub[0].strip()
                ws.cell(row=current_row, column=39).value = line_split_sub[1].strip()
            # MGW INDEX not working for SUB B
            elif "MGW INDEX" in line:
                line_split = line.split(':')
                line_split_sub = line_split[1].split('          ')
                ws.cell(row=current_row, column=30).value = line_split_sub[0].strip()
                ws.cell(row=current_row, column=43).value = line_split_sub[1].strip()
            # MGW NAME
            elif "MGW NAME" in line:
                line_split = line.split(':')
                line_split_sub = line_split[1].split('          ')
                ws.cell(row=current_row, column=31).value = line_split_sub[0].strip()
                ws.cell(row=current_row, column=44).value = line_split_sub[1].strip()
            elif "BNC CHAR" in line:
                line_split = line.split(':')
                line_split_sub = line_split[1].split('                      ')
                ws.cell(row=current_row, column=32).value = line_split_sub[0].strip()
                ws.cell(row=current_row, column=45).value = line_split_sub[1].strip()
            # TDM TERMID
            elif "TDM TERMID" in line:
                line_split = line.split(':')
            # NORMAL CTX ID
            elif "NORMAL CTX ID" in line:
                line_split = line.split(':')
                line_split_sub = line_split[1].split('                 ')
                ws.cell(row=current_row, column=34).value = line_split_sub[0].strip()
                ws.cell(row=current_row, column=47).value = line_split_sub[1].strip()
            # USED CODEC
            elif "USED CODEC" in line:
                line_split = line.split(':')
                line_split_sub = line_split[1].split('                ')
                ws.cell(row=current_row, column=35).value = line_split_sub[0].strip()
                ws.cell(row=current_row, column=48).value = line_split_sub[1].strip()
        # MSC SMS OBSERVATION REPORT

filename_S = sys.argv[1].split('.')[0] + "_" + sys.argv[2]
wb.save(filename_S + ".xlsm")
